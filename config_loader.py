"""
肉鸽游戏 - Excel 配置加载器
从 config.xlsx 读取配置并覆盖 settings 中的常量。
若 config.xlsx 不存在，自动生成默认模板。
用户可手动修改 Excel 并保存，下次启动游戏自动生效。
"""
import os
from openpyxl import load_workbook, Workbook

CONFIG_FILE = os.path.join(os.path.dirname(__file__), "config.xlsx")

# 配置类型映射：sheet名 -> (参数名, 类型, 中文备注)
# 类型: int / float / bool / str / tuple3 / tier
DEFAULT_CONFIG = {
    "通用表": {
        "FPS": ("int", 60, "游戏帧率（每秒刷新次数，一般 60）"),
        "SCREEN_WIDTH": ("int", 1280, "窗口宽度（像素）"),
        "SCREEN_HEIGHT": ("int", 720, "窗口高度（像素）"),
        "TILE_SIZE": ("int", 80, "地图一格的大小（像素，1格=80px）"),
        "GAME_TITLE": ("str", "肉鸽游戏 Demo", "窗口标题文字"),
        "SAVE_FILE": ("str", "save.json", "存档文件名"),
    },
    "角色表": {
        "PLAYER_WIDTH": ("int", 40, "角色宽度（像素）"),
        "PLAYER_HEIGHT": ("int", 60, "角色高度（像素）"),
        "PLAYER_SPEED": ("int", 300, "角色移动速度（像素/秒）"),
        "PLAYER_HP": ("int", 10, "角色初始血量"),
        "PLAYER_COLOR": ("tuple3", "90,180,255", "角色颜色（红,绿,蓝，0-255）"),
        "DASH_DISTANCE": ("int", 3, "冲刺位移距离（格，1格=80px）"),
        "DASH_DOUBLE_TAP_MS": ("int", 250, "双击方向键触发冲刺的时间（毫秒）"),
    },
    "敌人表": {
        "ENEMY_COUNT": ("int", 10, "初始生成的普通敌人数"),
        "ENEMY_SIZE": ("int", 32, "普通敌人基础直径（像素）"),
        "ENEMY_SPEED_FACTOR": ("float", 0.8, "普通敌人移速 = 角色移速 × 此倍数"),
        "ENEMY_HP_MIN": ("int", 1, "敌人血量最小值"),
        "ENEMY_HP_MAX": ("int", 5, "敌人血量最大值"),
        "ENEMY_SPAWN_INTERVAL": ("float", 1.0, "普通敌人生成间隔（秒，越小越快）"),
        "ENEMY_SPAWN_DIST_MIN": ("int", 1200, "敌人生成位置距玩家最小距离（像素，画面外）"),
        "ENEMY_SPAWN_DIST_MAX": ("int", 2500, "敌人生成位置距玩家最大距离（像素）"),
        "ENEMY_EMPTY_REFRESH": ("float", 1.0, "画面内无敌人多少秒后立即刷新敌人"),
        "ENEMY_TIER_STATS": ("tier", "1.0,1.0,1.0,1.0|1.5,2.0,1.5,1.2|2.0,2.5,2.0,1.4|2.5,3.0,2.5,1.6|3.0,3.5,3.0,1.8|3.5,4.0,3.5,2.0", "6个等级敌人属性（攻击,生命,速度,尺寸），每级用|分隔"),
        "ELITE_SPAWN_INTERVAL": ("int", 60, "精英怪生成间隔（秒，每 N 秒 1 个）"),
        "ELITE_SIZE_MULT": ("float", 2.0, "精英怪体型 = 普通敌人 × 此倍数"),
        "ELITE_HP_MULT": ("int", 10, "精英怪血量 = 普通敌人 × 此倍数"),
        "ELITE_ATK_MULT": ("int", 2, "精英怪攻击 = 普通敌人 × 此倍数"),
        "ELITE_SPD_MULT": ("float", 1.1, "精英怪移速 = 普通敌人 × 此倍数"),
        "ELITE_COLOR": ("tuple3", "255,200,60", "精英怪颜色（红,绿,蓝）"),
        "ARCHER_SIZE": ("int", 30, "弓箭手尺寸（像素）"),
        "ARCHER_HP": ("int", 3, "弓箭手血量"),
        "ARCHER_SPEED": ("int", 150, "弓箭手移速（像素/秒，比角色慢）"),
        "ARCHER_ATTACK_RANGE": ("int", 500, "弓箭手攻击距离（在此范围内保持距离射击）"),
        "ARCHER_FLEE_RANGE": ("int", 200, "玩家靠近到这个距离时弓箭手逃跑"),
        "ARCHER_FIRE_INTERVAL": ("float", 2.0, "弓箭手射击间隔（秒）"),
        "ARCHER_ARROW_DAMAGE": ("int", 1, "箭矢命中角色伤害"),
        "ARCHER_ARROW_SPEED": ("int", 400, "箭矢飞行速度（像素/秒）"),
        "ARCHER_COLOR": ("tuple3", "150,200,120", "弓箭手颜色（红,绿,蓝）"),
        "ARCHER_ARROW_COLOR": ("tuple3", "200,200,150", "箭矢颜色（红,绿,蓝）"),
    },
    "召唤表": {
        "SUMMON_INTERVAL": ("int", 10, "每多少秒召唤一只生肖"),
        "SUMMON_DURATION": ("int", 20, "召唤物持续秒数"),
        "SUMMON_SPEED": ("int", 240, "召唤物移动速度（像素/秒）"),
        "SUMMON_SIZE": ("int", 36, "召唤物尺寸（像素）"),
        "SUMMON_ATTACK": ("int", 1, "召唤物碰撞伤害"),
        "SUMMON_KNOCKBACK": ("int", 30, "召唤物击中敌人击退距离（像素）"),
        "SUMMON_ATTACK_SPD_MIN": ("float", 0.10, "召唤物攻速升级最低加成"),
        "SUMMON_ATTACK_SPD_MAX": ("float", 0.50, "召唤物攻速升级最高加成"),
        "SUMMON_MOVE_SPD_MIN": ("float", 0.10, "召唤物移速升级最低加成"),
        "SUMMON_MOVE_SPD_MAX": ("float", 0.20, "召唤物移速升级最高加成"),
        "SUMMON_DURATION_MIN": ("float", 0.10, "召唤物时长升级最低加成"),
        "SUMMON_DURATION_MAX": ("float", 0.20, "召唤物时长升级最高加成"),
        "SUMMON_ZODIAC": ("list", "鼠,牛,虎,兔,龙,蛇,马,羊,猴,鸡,狗,猪", "十二生肖顺序（逗号分隔）"),
        "SUMMON_SHOOTERS": ("list", "鼠,牛,虎,兔,龙,蛇", "射击型召唤物（逗号分隔）"),
        "SUMMON_BULLET_COLOR_MOUSE": ("tuple3", "200,200,220", "鼠子弹颜色"),
        "SUMMON_BULLET_COLOR_OX": ("tuple3", "180,140,100", "牛子弹颜色"),
        "SUMMON_BULLET_COLOR_TIGER": ("tuple3", "255,170,40", "虎子弹颜色"),
        "SUMMON_BULLET_COLOR_RABBIT": ("tuple3", "255,200,230", "兔子弹颜色"),
        "SUMMON_BULLET_COLOR_DRAGON": ("tuple3", "80,220,140", "龙子弹颜色"),
        "SUMMON_BULLET_COLOR_SNAKE": ("tuple3", "120,200,80", "蛇子弹颜色"),
    },
    "升级项表": {
        "KILLS_PER_LEVEL": ("int", 5, "每击杀多少敌人升一级"),
        "PLAYER_HP_BONUS": ("int", 2, "选【生命】时增加的血量上限"),
        "SPEED_BONUS_PCT": ("float", 0.10, "选【移速】时增加的移速比例（0.10=+10%）"),
        "ATTACK_SPEED_MIN": ("float", 0.10, "【攻速】升级最低加成（0.10=+10%）"),
        "ATTACK_SPEED_MAX": ("float", 1.00, "【攻速】升级最高加成（1.00=+100%）"),
        "BULLET_COUNT_MIN": ("int", 1, "【弹量】升级最少增加子弹数"),
        "BULLET_COUNT_MAX": ("int", 5, "【弹量】升级最多增加子弹数"),
        "PIERCE_BONUS": ("int", 1, "【穿透】每次增加穿透次数"),
        "BULLET_SIZE_BONUS_MIN": ("float", 0.50, "【子弹大小】最低加成比例"),
        "BULLET_SIZE_BONUS_MAX": ("float", 1.00, "【子弹大小】最高加成比例"),
        "BOUNCE_COUNT": ("int", 1, "【弹射】每次增加弹射次数"),
        "LIFESTEAL_CHANCE": ("float", 0.10, "【吸血】每次获得的吸血概率（0.10=+10%）"),
        "LIFESTEAL_PITY": ("int", 10, "吸血假概率保底（每多少发必触发1次）"),
        "LIFESTEAL_HEAL": ("int", 1, "每次吸血回血量"),
        "CRIT_RATE_DEFAULT": ("float", 0.05, "角色默认暴击率（0.05=5%）"),
        "CRIT_DAMAGE_DEFAULT": ("float", 2.0, "角色默认暴击伤害倍数（2.0=200%）"),
        "CRIT_RATE_MIN": ("float", 0.05, "【暴击率】升级最低加成"),
        "CRIT_RATE_MAX": ("float", 0.20, "【暴击率】升级最高加成"),
        "CRIT_DMG_MIN": ("float", 0.05, "【暴击伤害】升级最低加成"),
        "CRIT_DMG_MAX": ("float", 0.20, "【暴击伤害】升级最高加成"),
        "BRAINWASH_CHANCE": ("float", 0.0, "初始洗脑概率（0=无，需升级获得）"),
        "BRAINWASH_DURATION": ("float", 30.0, "洗脑出的友军持续秒数"),
        "BARRIER_RADIUS": ("int", 1, "【屏障】初始范围（格）"),
        "BARRIER_RADIUS_BONUS": ("float", 0.5, "【屏障范围】每次增加格数"),
        "BARRIER_FREQ_BONUS": ("float", 1.0, "【屏障频率】每次增加倍率"),
        "BARRIER_DAMAGE": ("int", 1, "屏障每秒对范围内敌人伤害"),
        "REGEN_AMOUNT": ("int", 1, "【生命回复】每秒回血量"),
        "EXPLOSION_RADIUS": ("int", 1, "【爆炸】初始爆炸范围（格）"),
        "EXPLOSION_RADIUS_BONUS": ("float", 0.3, "【爆炸】每次增加范围格数"),
        "EXPLOSION_STUN": ("float", 0.2, "爆炸时敌人停顿秒数"),
        "KNOCKBACK_DISTANCE": ("int", 1, "【击退】初始击退距离（格）"),
        "KNOCKBACK_BONUS": ("int", 1, "【击退】每次增加距离格数"),
        "LOW_HP_THRESHOLD": ("float", 0.10, "血量低于此比例时触发红屏警告（0.10=10%）"),
    },
        "AI设置表": {
        "AI_API_URL": ("str", "http://192.168.5.16:20128/v1", "大模型API地址"),
        "AI_API_KEY": ("str", "sk-b6f4d3879cc4a442-aroz1n-d4d3d242", "API密钥"),
        "AI_MODEL": ("str", "ds/deepseek-v4-flash", "模型名称"),
        "AI_ENABLED": ("bool", "1", "是否启用AI功能（1开/0关）"),
        "AI_TIMEOUT": ("int", 20, "AI请求超时秒数"),
    },
    "战斗表": {
        "BULLET_SIZE": ("int", 8, "子弹基础直径（像素）"),
        "BULLET_SPEED_FACTOR": ("int", 3, "子弹速度 = 角色移速 × 此倍数"),
        "BULLET_RANGE_FACTOR": ("int", 10, "子弹射程 = 角色移速 × 此倍数"),
        "BULLET_DAMAGE": ("int", 1, "子弹基础伤害"),
        "BULLET_COLOR": ("tuple3", "255,220,90", "子弹颜色（红,绿,蓝）"),
        "FIRE_RATE": ("float", 2.0, "基础射击频率（发/秒）"),
        "AUTO_ATTACK_RANGE": ("int", 600, "自动攻击距离（敌人进入此范围自动开火）"),
        "DAMAGE_TEXT_LIFETIME": ("float", 0.5, "伤害跳字显示秒数"),
        "DAMAGE_TEXT_COLOR": ("tuple3", "255,255,255", "普通伤害跳字颜色"),
        "CRIT_TEXT_COLOR": ("tuple3", "255,60,60", "暴击伤害跳字颜色"),
        "EXPLOSION_COLOR": ("tuple3", "255,140,30", "爆炸特效颜色（红,绿,蓝）"),
    },
}


def _parse_value(type_str, raw):
    """按类型解析 Excel 单元格值"""
    if raw is None:
        return None
    if type_str == "int":
        return int(float(raw))
    elif type_str == "float":
        return float(raw)
    elif type_str == "bool":
        if isinstance(raw, str):
            return raw.strip().lower() in ("1", "true", "是", "yes")
        return bool(raw)
    elif type_str == "str":
        return str(raw)
    elif type_str == "tuple3":
        parts = str(raw).split(",")
        return (int(float(parts[0].strip())), int(float(parts[1].strip())), int(float(parts[2].strip())))
    elif type_str == "tier":
        # 每行用 | 分隔，每项用 , 分隔
        rows = []
        for line in str(raw).split("|"):
            parts = [float(x.strip()) for x in line.split(",") if x.strip()]
            rows.append(tuple(parts))
        return rows
    elif type_str == "list":
        # 逗号分隔的字符串列表（如 鼠,牛,虎）
        return [x.strip() for x in str(raw).split(",") if x.strip()]
    return raw


def _generate_default():
    """生成默认 Excel 模板"""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, items in DEFAULT_CONFIG.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["参数名", "值", "类型", "中文备注"])
        for name, (type_str, default, desc) in items.items():
            ws.append([name, default, type_str, desc])
        # 设置列宽
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 55
    wb.save(CONFIG_FILE)
    return wb


def load_config():
    """
    加载 config.xlsx 配置，返回 {参数名: 值} 字典。
    若文件不存在则生成默认模板。
    """
    if not os.path.exists(CONFIG_FILE):
        _generate_default()

    config = {}
    try:
        wb = load_workbook(CONFIG_FILE, data_only=True)
        for sheet_name, items in DEFAULT_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None or len(row) < 2:
                    continue
                param = row[0]
                value = row[1]
                if param is None:
                    continue
                # 从 DEFAULT_CONFIG 找类型
                if param in items:
                    t = items[param][0]
                else:
                    t = "str"
                parsed = _parse_value(t, value)
                if parsed is not None:
                    config[param] = parsed
    except Exception as e:
        print(f"[配置] 读取 config.xlsx 失败: {e}，使用默认值")
    return config


def get_config():
    """返回配置字典（供 settings.py 使用）"""
    return load_config()


if __name__ == "__main__":
    cfg = load_config()
    print(f"已加载 {len(cfg)} 个配置项")
    print("角色表关键项:", {k: cfg.get(k) for k in ["PLAYER_SPEED", "PLAYER_HP"]})
    print("敌人表关键项:", {k: cfg.get(k) for k in ["ENEMY_SPAWN_INTERVAL", "ELITE_HP_MULT"]})
    print("升级项表关键项:", {k: cfg.get(k) for k in ["KILLS_PER_LEVEL", "CRIT_RATE_DEFAULT"]})
